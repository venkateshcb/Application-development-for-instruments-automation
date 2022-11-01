from tkinter import *               #library for GUI operation 
from tkinter import messagebox      #library for pop-up operation
import serial.tools.list_ports
import serial                       #library for serial communication operation using "UART"
from openpyxl import load_workbook  # importing library for excel file operation
import pyvisa                       #importing library for virtual instruments software application
import time                         #importing library for dealy operation

class RootGUI():
    def __init__(self):
        '''Initializing the root GUI and other comps of the program'''
        self.root = Tk()
        self.root.title("Serial communication")
        self.root.geometry("900x550")
        self.root.config(bg="#C2C2C2")
        self.root.minsize(width = 1000, height = 650)#-----|----> min and max size of the GUI window
        self.root.maxsize(width = 1000, height = 650)#-----|

# Class to setup and create the communication manager
class ComGui():
    def __init__(self, root, serial):
        '''
        Initialize the connexion GUI and initialize the main widgets 
        '''
        self.rm = pyvisa.ResourceManager()
        self.inst = self.rm.open_resource('GPIB0::28::INSTR') #signal generator     GPIB address
        self.inst1 = self.rm.open_resource('GPIB0::20::INSTR') #spectrum analyser   GPIB address

        self.path_entry = Entry(root) #---------|----> seeting entry menu for path entry
        self.path_entry.config(width = 55)#-----|


        self.data = {'2': 0, '3': 0, '5': 0, '6': 0, '8': 0, '9': 0, '10': 0, 
        '11': 0, '12': 0, '13': 0, '21': 0, '23': 0, '24': 0, '26': 0, '27': 0, '28': 0, '30': 0, '31': 0}
        
        self.disp = f"BA1 : {self.data['28']}   BA2 : {self.data['10']}"    #display for bite attenuation
        self.a = 0
        self.b = 0
        self.c = 0
        # Initializing the Widgets
        self.root = root
        self.serial = serial
        
        self.frame = LabelFrame(root, text="Com Manager",
                                padx=5, pady=5, bg="#C2C2C2")
        
        self.frame1 = LabelFrame(root, text="TX_RX",
                                padx=5, pady=5, bg="#C2C2C2")
        
        self.frame2 = LabelFrame(root, text="ATTENUATION",
                                padx=5, pady=5, bg="#C2C2C2")

        self.frame3 = LabelFrame(root, text="PHASE",
                                padx=5, pady=5, bg="#C2C2C2")
        
        self.frame4 = LabelFrame(root, text="DDS_SWEEP1_SYNTH",
                                padx=5, pady=5, bg="#C2C2C2")
        
        self.frame5 = LabelFrame(root, text="1st UPCONVERTER",
                               padx=5, pady=5, bg="#C2C2C2")
    
        self.frame6 = LabelFrame(root, text="RFBITE_RXCAL",
                               padx=5, pady=5, bg="#C2C2C2")
    
        self.frame7 = LabelFrame(root, text="RF_FIELD",
                                padx=5, pady=5, bg="#C2C2C2")
        
        self.frame8 = LabelFrame(root, text="TXD\TXCAL",
                                padx=5, pady=5, bg="#C2C2C2")
        self.frame9 = LabelFrame(root, text="RX_BITE_1",
                                padx=5, pady=5, bg="#C2C2C2")
        self.frame10 = LabelFrame(root, text="RXCAL",padx=5,
                                pady=5, bg="#C2C2C2")
        self.frame11 = LabelFrame(root, text="LO1",
                                padx=5, pady=5, bg="#C2C2C2")
        self.frame12 = LabelFrame(root, text="RX_BITE_2",
                                padx=5, pady=5, bg="#C2C2C2")
        
        self.label_com = Label(
            self.frame, text="Available Port(s): ", bg="#C2C2C2", width=15, anchor="w")
        self.label_bd = Label(
            self.frame, text="Baude Rate: ", bg="#C2C2C2", width=15, anchor="w")
        
        self.label_tx = Label(
            self.frame1, text="Tx_Rx: ", bg="#C2C2C2", width=15, anchor="w")
        
        self.label_bite = Label(
            self.frame2, text="BITE_ATT: ", bg="#C2C2C2", width=15, anchor="w")
        self.label_txbite = Label(
            self.frame2, text="TXBITE_ATT: ", bg="#C2C2C2", width=15, anchor="w")
        self.dis_bite_att = Label(
            self.frame2, text=self.disp, bg="#C2C2C2", width=15, anchor="w")

        self.label_phase = Label(
            self.frame3, text="Phase in degree: ", bg="#C2C2C2", width=15, anchor="w")

        self.label_BPF_cTL1 = Label(
            self.frame4, text="BPF_CTL1: ", bg="#C2C2C2", width=15, anchor="w")
        self.label_BPF_cTL2 = Label(
            self.frame5, text="BPF_CTL2: ", bg="#C2C2C2", width=15, anchor="w")
        
        self.label_RX_CAL = Label(
            self.frame6, text="RFBITE_RXCAL: ", bg="#C2C2C2", width=15, anchor="w")
        
        self.label_Rf = Label(
            self.frame7, text="RF_FIELD: ", bg="#C2C2C2", width=15, anchor="w")

        self.label_txd = Label(
            self.frame8, text="Port : J16 \n Frequency :  9-10 GHz \n Power Level : +14dBm ± 1.5dB",
             bg="#C2C2C2", width=28, anchor="w")    
        self.label_Rx_bite = Label(
            self.frame9, text="Port : J8 \n Frequency :  9-10 GHz \n Power Level : -50dBm ± 2dB", 
            bg="#C2C2C2", width=28, anchor="w")
        self.label_Rxcal = Label(
            self.frame10, text="Port : J19 \n Frequency :  9-10 GHz \n Power Level : -26dBm ± 1dB", 
            bg="#C2C2C2", width=28, anchor="w")
        self.label_lo1 = Label(
            self.frame11, text="Port : J14 \n Frequency :  7.3 - 8.38 GHz \n Power Level : -14dBm ± 1.5dB", 
            bg="#C2C2C2", width=28, anchor="w")
        self.label_rx_bite_2 = Label(
            self.frame12, text="Port : J9 \n Frequency :  9 - 10 GHz \n Power Level : -50dBm ± 2dB", 
            bg="#C2C2C2", width=28, anchor="w")

        self.txd_button = Button(self.frame8,text = "measure" , 
        width = 8, height = 1, command = lambda : [self.cal_bite(), self.calculate(), self.cal_txbite(), self.txd_txcal()])
        self.rx_bite_button = Button(self.frame9,text = "measure" , 
        width = 8, height = 1, command = lambda : [self.cal_bite(), self.calculate(), self.cal_txbite(), self.rx_bite1()])
        self.rxcal_button= Button(self.frame10,text = "measure" , 
        width = 8, height = 1, command = lambda : [self.cal_bite(), self.calculate(), self.cal_txbite(), self.rxcal()])
        self.lo1_button = Button(self.frame11,text = "measure" ,
         width = 8, height = 1, command = lambda : [self.cal_bite(), self.calculate(), self.cal_txbite(), self.lo1_output()])
        self.rx_bite_button_2 = Button(self.frame12,text = "measure" ,
         width = 8, height = 1, command = lambda : [self.cal_bite(), self.calculate(), self.cal_txbite(), self.rx_bite2()])

        # Setup the Drop option menu
        self.baudOptionMenu()#--------------|
        self.ComOptionMenu() #--------------|
        self.tx_rx()#-----------------------|
        self.attenuat() #-------------------|
        self.phas()#------------------------|------------> initialising the all required functions while staring the code
        self.dds_sweep()#-------------------|
        self.upconv()#----------------------|
        self.rx_cal()#----------------------|
        self.rf_field()#--------------------|

        # Add the control buttons for refreshing the COMs & Connect
        self.btn_refresh = Button(self.frame, text="Refresh",
                                  width=10,  command=self.com_refresh)
        self.btn_connect = Button(self.frame, text="Connect",
                                  width=10, state="disabled",  command=self.serial_connect)
        
        self.phase.but = Button(root,text = "Send" , bg = "gray", width = 10, height = 1, 
        command = lambda : [self.cal_bite(), self.calculate(), self.cal_txbite(), self.send_data()])

        # Optional Graphic parameters
        self.padx = 20
        self.pady = 5

        # Put on the grid all the elements
        self.publish()

    def publish(self):
        '''
         Method to display all the Widget of the main frame
        '''
        self.frame.grid(row=0, column=0, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame1.grid(row=0, column=4, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame2.grid(row=0, column=8, rowspan=3,
                        columnspan=3, padx=5, pady=5)

        self.frame3.grid(row=4, column=0, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame4.grid(row=4, column=4, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame5.grid(row=4, column=8, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        
        self.frame6.grid(row=8, column=0, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame7.grid(row=8, column=4, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame8.grid(row=20, column=0, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame9.grid(row=20, column=4, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame10.grid(row=20, column=8, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame11.grid(row=23, column=0, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        self.frame12.grid(row=23, column=4, rowspan=3,
                        columnspan=3, padx=5, pady=5)
        
        self.path_entry.grid(column = 4, row = 18)
        self.label_com.grid(column=1, row=2)
        self.label_bd.grid(column=1, row=3)
        self.com_label.grid(column = 9, row = 8)
        

        self.drop_baud.grid(column=2, row=3, padx=self.padx, pady=self.pady)
        self.drop_com.grid(column=2, row=2, padx=self.padx)

        self.btn_refresh.grid(column=3, row=2)
        self.btn_connect.grid(column=3, row=3)

        self.txd_button.grid(column = 4, row = 20)
        self.rx_bite_button.grid(column = 8, row = 20)
        self.rxcal_button.grid(column = 8, row = 20)
        self.lo1_button.grid(column = 0, row = 23)
        self.rx_bite_button_2.grid(column = 0, row = 23)

        self.drop_tx.grid(column = 5, row = 1, padx=self.padx, pady=self.pady)
        self.label_tx.grid(column=5, row=0)
        
        self.label_bite.grid(column = 7, row = 1, padx=self.padx, pady=self.pady)
        self.beatt.grid(column = 7, row = 3)
        self.dis_bite_att.grid(column = 7, row = 5)

        self.label_txbite.grid(column = 10, row = 1)
        self.eatt.grid(column = 10, row = 3)
        
        self.label_phase.grid(column = 0, row = 4, padx=self.padx, pady=self.pady)
        self.phase.grid(column =0, row = 6, padx=self.padx, pady=self.pady)
        self.phase.but.grid(column =9, row = 9, padx=self.padx, pady=self.pady)
        
        self.label_BPF_cTL1.grid(column = 4, row = 5,padx=self.padx, pady=self.pady)
        self.drop_dds.grid(column = 4, row = 6, pady=self.pady)
        
        self.label_BPF_cTL2.grid(column = 5, row = 5,padx=self.padx, pady=self.pady)
        self.drop_upconv.grid(column = 5, row = 6,  pady=self.pady)

        self.label_RX_CAL.grid(column= 0, row = 8, padx=self.padx, pady=self.pady)
        self.drop_rx_cal.grid(column = 0, row = 9, padx=self.padx, pady=self.pady)

        self.drop_rf.grid(column = 4, row = 9, padx=self.padx, pady=self.pady)
        self.label_Rf.grid(column = 4, row = 8, padx=self.padx, pady=self.pady)
        self.label_txd.grid(column = 4, row = 8, padx=self.padx, pady=self.pady)
        self.label_Rx_bite.grid(column = 8, row = 8, padx=self.padx, pady=self.pady)
        self.label_Rxcal.grid(column = 8, row = 8, padx=self.padx, pady=self.pady)
        self.label_lo1.grid(column = 0, row = 22, padx=self.padx, pady=self.pady)
        self.label_rx_bite_2.grid(column = 0, row = 22, padx=self.padx, pady=self.pady)

    def ComOptionMenu(self):
        '''
         Method to Get the available COMs connected to the PC
         and list them into the drop menu
        '''
        # Generate the list of available coms
        port = serial.tools.list_ports.comports(include_links=False)
        coms = [] 
        self.s = ""
        for i in port:
            coms.append(i.device)
            self.s += f"{i} \n"
        self.clicked_com = StringVar()
        self.clicked_com.set(coms[0])
        self.drop_com = OptionMenu(
            self.frame, self.clicked_com, *coms, command=self.connect_ctrl)
        
        self.com_label = Label(self.root, text = self.s , bg="#C2C2C2", width = 43)
        self.drop_com.config(width=10)

    def baudOptionMenu(self):
        '''
         Method to list all the baud rates in a drop menu
        '''
        self.clicked_bd = StringVar()
        bds = ["9600"]
        self.clicked_bd .set(bds[0])
        self.drop_baud = OptionMenu(
            self.frame, self.clicked_bd, *bds, command=self.connect_ctrl)
        self.drop_baud.config(width=10)

    def tx_rx(self):
        self.clicked_tx = StringVar()
        tx = ['0', '1']
        self.clicked_tx.set(tx[0])
        self.drop_tx = OptionMenu(self.frame1, self.clicked_tx, *tx)
        self.drop_tx.config(width = 10)
    
    def attenuat(self):
        #txBITE ATTENUATION CONTROLS
        self.eatt = Entry(self.frame2)
        self.eatt.config(width = 10)
        
        self.beatt = Entry(self.frame2)
        self.beatt.config(width = 10)

    def phas(self):
        self.phase = Entry(self.frame3)
        self.phase.config(width = 15)

    def dds_sweep(self):
        self.clicked_dds = StringVar()
        dds = ['0', '1']
        self.clicked_dds.set(dds[0])
        self.drop_dds = OptionMenu(self.frame4, self.clicked_dds, *dds)
        self.data['31'] = self.clicked_dds.get()
        self.drop_dds.config(width = 10)

    def cal(self, data, step, datalen):
        # common function to convert decimal to binary and setting the bits
        f = list(bin(int(data/step)))[2::]
        f = f[::-1]
        if len(f) < datalen:
            for i in range(datalen-len(f)):
                f.append("0")
        return f

    def calculate(self):
        s = self.phase.get()
        #exception method is used to try to get the values and if 
        #values are absent it will show error message to enter the value or to entre the correct value
        try:
            if (s != "") & (float(s) % 22.5 == 0) & (float(s) >= 0) & (float(s) < 337.6):
                s = float(s)
                l1 = self.cal(s, 22.5, 4)
                self.data['23'] = l1[0]
                self.data['5'] = l1[1]
                self.data['24'] = l1[2]
                self.data['6'] = l1[3]
                self.a = 1
            else:
                messagebox.showerror("PHASE", "Enter the value in step of 22.5ͦ  or less than or equal to 337.5")
        except:
            messagebox.showerror("PHASE", "Enter the value in step of 22.5ͦ  or less than or equal to 337.5")

    def cal_txbite(self):
        d = self.eatt.get()
        try:
            if (d != "") & (float(d) % 0.5 == 0) & (float(d) >= 0) & (float(d) < 3.6):
                d = float(d)
                l2 = self.cal(d, 0.5, 3)
                self.data['2'] = l2[0]
                self.data['21'] = l2[1]
                self.data['3'] = l2[2]
                self.b = 1
            else:
                messagebox.showerror("ATTENUATION", "Enter the values in step of 0.5 and less than or equal to 3.5")
        except:
            messagebox.showerror("ATTENUATION", "Enter the values in step of 0.5 and less than or equal to 3.5")

    def cal_bite(self):
        q = self.beatt.get()
        try:
            if (q != "")&(int(q) % 1 == 0) & (int(q) >= 0) & (int(q) < 64):
                q = int(q)
                l3 = self.cal(q, 1, 6)
                self.data['26'] = l3[0]
                self.data['8'] = l3[1]
                self.data['27'] = l3[2]
                self.data['9'] = l3[3]
                self.data['28'] = l3[4]
                self.data['10'] = l3[5]
                self.c = 1
                if q < 32:
                    x = q
                    y = 0
                else:
                    x = q-32
                    y = 32
                self.disp = f"BA1 : {x}   BA2 : {y}"
                self.dis_bite_att.destroy()
                self.dis_bite_att = Label(
                    self.frame2, text=self.disp, bg="#C2C2C2", width=15, anchor="w")
                self.dis_bite_att.grid(column = 7, row = 5)
            else:
                messagebox.showerror("ATTENUATION", "Enter the integer values less than or Equal to 63")
        except:
            messagebox.showerror("ATTENUATION", "Enter the integer values less than or Equal to 63")

    def upconv(self):
        self.clicked_upconv = StringVar()
        up = ['0', '1']
        self.clicked_upconv.set(up[0])
        self.drop_upconv = OptionMenu(self.frame5, self.clicked_upconv, *up)
        self.drop_upconv.config(width = 10)
    
    def rx_cal(self):
        self.clicked_rx_cal= StringVar()
        rx_cal = ['0', '1']
        self.clicked_rx_cal.set(rx_cal[0])
        self.drop_rx_cal = OptionMenu(self.frame6, self.clicked_rx_cal, *rx_cal)
        self.drop_rx_cal.config(width = 10)

    def rf_field(self):
        self.clicked_rf= StringVar()
        rf = ['0', '1']
        self.clicked_rf.set(rf[0])
        self.drop_rf = OptionMenu(self.frame7, self.clicked_rf, *rf)
        self.data['30'] = self.clicked_rf.get()
        self.drop_rf.config(width = 10)

    def txd_txcal(self):
        self.path = self.path_entry.get()
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self.ws1 = self.wb['Cable loss']
        c = 4       
        self.inst.write(f":OUTP ON") #for power level
        self.inst.write(f":SOURce1:POWer:LEVel:IMMediate:AMPLitude 6") #for power level
        self.data['11'] = self.clicked_rx_cal.get()
        self.data['12'] = self.clicked_tx.get()
        self.data['30'] = self.clicked_rf.get()
        for i, j , k in zip(self.ws["D4:D32"], self.ws1["C4:C32"], self.ws["E4:E32"]):            
            for cell1, cell2, cell3 in zip(i, j, k):
                if (cell1.value >= 420) & (cell1.value <= 595):
                    self.data['31'] = "0"
                else:
                    self.data['31'] = "1"
                if (cell3.value == 4.92):
                    self.data['13'] = "0"
                else:
                    self.data['13'] = "1"
                p = []
                for i in self.data.values():
                    p.append(str(i))       
                p = ''.join(p)
                p = [int(p[0:8], 2), int(p[8:16], 2), int(p[16:18], 2)]
        
                PORT = self.clicked_com.get()
                BAUD = self.clicked_bd.get()
                self.ser = serial.Serial()
                self.ser.baudrate = BAUD
                self.ser.port = PORT
                self.ser.open()
                self.ser.write(bytearray(p))
                self.ser.close()
                self.inst.write(f":SOURce1:FREQuency:CW {int(cell1.value)*10**6}") #for frequency adjustment
                time.sleep(1) #delay of one second
                self.inst1.write("CALC:MARK1 ON")   #sets the marker 1 on
                self.inst1.write(":CALC1:MARK1:MAX:PEAK")
                a = self.inst1.query(":CALC:MARK1:Y?") #getting power level
                b = self.inst1.query(":CALC:MARK1:X?") #getting frequency level
                a = round(float(a),2)
                b = int(b)
                self.ws[f"H{c}"] = a + float(cell2.value)*(-1)
                self.ws[f"G{c}"] = b
                c += 1       
        self.wb.save(self.path)
    
    def lo1_output(self):
        self.path = self.path_entry.get()
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self.ws1 = self.wb['Cable loss']
        c = 4       
        self.inst.write(f":OUTP ON") #for power level
        self.inst.write(f":SOURce1:POWer:LEVel:IMMediate:AMPLitude 6") #for power level
        self.data['11'] = self.clicked_rx_cal.get()
        self.data['12'] = self.clicked_tx.get()
        self.data['30'] = self.clicked_rf.get()
        for i, j , k in zip(self.ws["D4:D32"], self.ws1["C4:C32"], self.ws["E4:E32"]):            
            for cell1, cell2, cell3 in zip(i, j, k):
                if (cell1.value >= 420) & (cell1.value <= 595):
                    self.data['31'] = "0"
                else:
                    self.data['31'] = "1"
                if (cell3.value == 4.92):
                    self.data['13'] = "0"
                else:
                    self.data['13'] = "1"
                p = []
                for i in self.data.values():
                    p.append(str(i))       
                p = ''.join(p)
                p = [int(p[0:8], 2), int(p[8:16], 2), int(p[16:18], 2)]
        
                PORT = self.clicked_com.get()
                BAUD = self.clicked_bd.get()
                self.ser = serial.Serial()
                self.ser.baudrate = BAUD
                self.ser.port = PORT
                self.ser.open()
                self.ser.write(bytearray(p))
                self.ser.close()
                self.inst.write(f":SOURce1:FREQuency:CW {int(cell1.value)*10**6}") #for frequency adjusting
                time.sleep(1)
                self.inst1.write("CALC:MARK1 ON")   #sets the marker 1 on
                self.inst1.write(":CALC1:MARK1:MAX:PEAK")
                a = self.inst1.query(":CALC:MARK1:Y?") #getting power level
                a = round(float(a),2)
                self.ws[f"I{c}"] = a + float(cell2.value)*(-1)
                c += 1       
        self.wb.save(self.path)
    
    def rx_bite1(self):
        self.path = self.path_entry.get()
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self.ws1 = self.wb['Cable loss']
        c = 4       
        self.inst.write(f":OUTP ON") #for power level
        self.inst.write(f":SOURce1:POWer:LEVel:IMMediate:AMPLitude 6") #for power level
        self.data['11'] = self.clicked_rx_cal.get()
        self.data['12'] = self.clicked_tx.get()
        self.data['30'] = self.clicked_rf.get()
        for i, j , k in zip(self.ws["D4:D32"], self.ws1["C4:C32"], self.ws["E4:E32"]):            
            for cell1, cell2, cell3 in zip(i, j, k):
                if (cell1.value >= 420) & (cell1.value <= 595):
                    self.data['31'] = "0"
                else:
                    self.data['31'] = "1"
                if (cell3.value == 4.92):
                    self.data['13'] = "0"
                else:
                    self.data['13'] = "1"
                p = []
                for i in self.data.values():
                    p.append(str(i))       
                p = ''.join(p)
                p = [int(p[0:8], 2), int(p[8:16], 2), int(p[16:18], 2)]
        
                PORT = self.clicked_com.get()
                BAUD = self.clicked_bd.get()
                self.ser = serial.Serial()
                self.ser.baudrate = BAUD
                self.ser.port = PORT
                self.ser.open()
                self.ser.write(bytearray(p))
                self.ser.close()
                self.inst.write(f":SOURce1:FREQuency:CW {int(cell1.value)*10**6}") #for frequency adjusting
                time.sleep(1)
                self.inst1.write("CALC:MARK1 ON")   #sets the marker 1 on
                self.inst1.write(":CALC1:MARK1:MAX:PEAK")
                a = self.inst1.query(":CALC:MARK1:Y?") #getting power level
                a = round(float(a),2)
                self.ws[f"K{c}"] = a + float(cell2.value)*(-1)
                c += 1       
        self.wb.save(self.path)
    
    def rx_bite2(self):
        self.path = self.path_entry.get()
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self.ws1 = self.wb['Cable loss']
        c = 4       
        self.inst.write(f":OUTP ON") #for power level
        self.inst.write(f":SOURce1:POWer:LEVel:IMMediate:AMPLitude 6") #for power level
        self.data['11'] = self.clicked_rx_cal.get()
        self.data['12'] = self.clicked_tx.get()
        self.data['30'] = self.clicked_rf.get()
        for i, j , k in zip(self.ws["D4:D32"], self.ws1["C4:C32"], self.ws["E4:E32"]):            
            for cell1, cell2, cell3 in zip(i, j, k):
                if (cell1.value >= 420) & (cell1.value <= 595):
                    self.data['31'] = "0"
                else:
                    self.data['31'] = "1"
                if (cell3.value == 4.92):
                    self.data['13'] = "0"
                else:
                    self.data['13'] = "1"
                p = []
                for i in self.data.values():
                    p.append(str(i))       
                p = ''.join(p)
                p = [int(p[0:8], 2), int(p[8:16], 2), int(p[16:18], 2)]
        
                PORT = self.clicked_com.get()
                BAUD = self.clicked_bd.get()
                self.ser = serial.Serial()
                self.ser.baudrate = BAUD
                self.ser.port = PORT
                self.ser.open()
                self.ser.write(bytearray(p))
                self.ser.close()
                self.inst.write(f":SOURce1:FREQuency:CW {int(cell1.value)*10**6}") #for frequency adjusting
                time.sleep(1)
                self.inst1.write("CALC:MARK1 ON")   #sets the marker 1 on
                self.inst1.write(":CALC1:MARK1:MAX:PEAK")
                a = self.inst1.query(":CALC:MARK1:Y?") #getting power level
                a = round(float(a),2)
                self.ws[f"L{c}"] = a + float(cell2.value)*(-1)
                c += 1       
        self.wb.save(self.path)

    def rxcal(self):
        self.path = self.path_entry.get()
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self.ws1 = self.wb['Cable loss']
        c = 4       
        self.inst.write(f":OUTP ON") #for power level
        self.inst.write(f":SOURce1:POWer:LEVel:IMMediate:AMPLitude 6") #for power level
        self.data['11'] = self.clicked_rx_cal.get()
        self.data['12'] = self.clicked_tx.get()
        self.data['30'] = self.clicked_rf.get()
        for i, j , k in zip(self.ws["D4:D32"], self.ws1["C4:C32"], self.ws["E4:E32"]):            
            for cell1, cell2, cell3 in zip(i, j, k):
                if (cell1.value >= 420) & (cell1.value <= 595):
                    self.data['31'] = "0"
                else:
                    self.data['31'] = "1"
                if (cell3.value == 4.92):
                    self.data['13'] = "0"
                else:
                    self.data['13'] = "1"
                p = []
                for i in self.data.values():
                    p.append(str(i))       
                p = ''.join(p)
                p = [int(p[0:8], 2), int(p[8:16], 2), int(p[16:18], 2)]
        
                PORT = self.clicked_com.get()
                BAUD = self.clicked_bd.get()
                self.ser = serial.Serial()
                self.ser.baudrate = BAUD
                self.ser.port = PORT
                self.ser.open()
                self.ser.write(bytearray(p))
                self.ser.close()
                self.inst.write(f":SOURce1:FREQuency:CW {int(cell1.value)*10**6}") #for frequency adjusting
                time.sleep(1)
                self.inst1.write("CALC:MARK1 ON")   #sets the marker 1 on
                self.inst1.write(":CALC1:MARK1:MAX:PEAK")
                a = self.inst1.query(":CALC:MARK1:Y?") #getting power level
                a = round(float(a),2)
                self.ws[f"J{c}"] = a + float(cell2.value)*(-1)
                c += 1       
        self.wb.save(self.path)

    def connect_ctrl(self, widget):
        '''
        Mehtod to keep the connect button disabled if all the 
        conditions are not cleared
        '''
        if "-" in self.clicked_bd.get() or "-" in self.clicked_com.get():
            self.btn_connect["state"] = "disabled"
        else:
            self.btn_connect["state"] = "active"

    def com_refresh(self):
        self.drop_com.destroy()
        self.com_label.destroy()
        self.ComOptionMenu()
        self.drop_com.grid(column=2,row=2, padx= self.padx)
        self.com_label.grid(column = 9, row = 8)    

    def serial_connect(self):
        if self.btn_connect["text"] in "Connect":
            # Start the serial communication
            self.serial.serialopen(self)

            # If connection established move on
            if self.serial.ser.status:
                # Update the COM manager
                self.btn_connect["text"] = "Disconnect"
                self.btn_refresh["state"] = "disable"
                self.drop_baud["state"] = "disable"
                self.drop_com["state"] = "disable"
                InfoMsg = f"Successful UART connection using {self.clicked_com.get()}"
                messagebox.showinfo("showinfo", InfoMsg)

            else:
                ErrorMsg = f"Failure to estabish UART connection using{self.clicked_com.get()} "
                messagebox.showerror("showerror", ErrorMsg)
        else:

            # Closing the Serial COM
            # Close the serial communication
            self.serial.serialclose(self)

            InfoMsg = f"UART connection using {self.clicked_com.get()} is now closed"
            messagebox.showwarning("showinfo", InfoMsg)
            self.btn_connect["text"] = "Connect"
            self.btn_refresh["state"] = "active"
            self.drop_baud["state"] = "active"
            self.drop_com["state"] = "active"

    def send_data(self):
        self.data['11'] = self.clicked_rx_cal.get()#--|
        self.data['12'] = self.clicked_tx.get()#------|
        self.data['13'] = self.clicked_upconv.get()#--|-->getting data from entry menu
        self.data['30'] = self.clicked_rf.get()#------|
        self.data['31'] = self.clicked_dds.get()#-----|
        p = []
        for i in self.data.values():
            p.append(str(i))       
        p = ''.join(p)  #convertes list to string
        p = [int(p[0:8], 2), int(p[8:16], 2), int(p[16:18], 2)] #conversion of string to integer of base 2 for converting it into hexa decimal number
        if (self.a == 1) & (self.b == 1) & (self.c == 1):   #cheks for condition to send data.
            PORT = self.clicked_com.get()
            BAUD = self.clicked_bd.get()
            self.ser = serial.Serial()
            self.ser.baudrate = BAUD
            self.ser.port = PORT
            self.ser.open()
            self.ser.write(bytearray(p))
            self.ser.close()
            self.a = 0
            self.b = 0
            self.c = 0
        else:
            messagebox.showerror("ERROR","Data is not sent")

class SerialCtrl:
    #checks for serial port to open the port
    def __init__(self):
        pass
    
    def serialopen(self, ComGui):
        try:
            self.ser.is_open
        except:
            PORT = ComGui.clicked_com.get()
            BAUD = ComGui.clicked_bd.get()
            self.ser = serial.Serial()
            self.ser.baudrate = BAUD
            self.ser.port = PORT
            self.ser.timeout = 0.1
        

        try:
            if self.ser.is_open:
                print("serial port --> already open")
                self.ser.status = True
            else:
                PORT = ComGui.clicked_com.get()
                BAUD = ComGui.clicked_bd.get()
                self.ser = serial.Serial()
                self.ser.baudrate = BAUD
                self.ser.port = PORT
                self.ser.timeout = 0.1
                #self.ser.open()
                self.ser.status = True
        
        except:
            self.ser.status = False
    
    def serialclose(self, ComGui):
        try:
            self.ser.is_open
            self.ser.close()
            self.ser.status = False
        except:
            self.ser.status = True


if __name__ == "__main__":
    MySerial = SerialCtrl()
    RootMaster = RootGUI()
    # Initiate the Communication Master class that will manage all the other GUI classes
    ComMaster = ComGui(RootMaster.root, MySerial)
    # Start the Graphic User Interface
    RootMaster.root.mainloop()